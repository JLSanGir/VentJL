from openpyxl import *

archivoexcel = "S:\Principal\partesproduccion.xlsm"
#archivoexcel = "C:\partesproduccion.xlsm"
#archivoexcel = "F:\DatosF\partesproduccion.xlsm"

def ver_sheets(namesheet, mmin_col, mmax_col):
    excel_doc = load_workbook(archivoexcel)
    sheet = excel_doc.get_sheet_by_name(namesheet)

    ultimafilahoja = sheet.max_row
    valor_col = mmax_col
    if sheet.max_column < mmax_col:
        valor_col = sheet.max_column

    all_rows = sheet.iter_rows(min_row=1, max_row=ultimafilahoja, min_col=mmin_col, max_col=valor_col,  values_only=True)
    return all_rows, ultimafilahoja

def ver_celda(namesheet,row, col):
    excel_doc = load_workbook(archivoexcel)
    sheet = excel_doc.get_sheet_by_name(namesheet)
    celda = sheet.cell(row, col)
    return celda.value

def rec_columna(namesheet, col):
    excel_doc = load_workbook(archivoexcel)
    sheet = excel_doc.get_sheet_by_name(namesheet)
    ultimafilahoja = sheet.max_row

    for i in range(col, ultimafilahoja):
        if not sheet.cell(row=i, column=col).value is None:
            cell_obj = sheet.cell(row=i, column=col)
            print(cell_obj.value)

    for r in range(1, ultimafilahoja):
        if not sheet.cell(row=r, column=3).value is None:
            ultimafila = r + 1
        else:
            break

# grab the active worksheet
#ws = wb.active

# Data can be assigned directly to cells
#ws['A1'] = 42

# Rows can also be appended
#ws.append([1, 2, 3])

# Python types will automatically be converted
#import datetime
#ws['A2'] = datetime.datetime.now()

# Save the file
#wb.save("sample.xlsx")